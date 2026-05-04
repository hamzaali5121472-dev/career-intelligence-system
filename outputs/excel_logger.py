"""
Excel Logger — B2Have Career Intelligence System
Maintains a rolling master log of every article/post processed by the pipeline.

File: data/intelligence_log.xlsx
Every pipeline run appends new rows (deduplicated by URL).
Existing rows are never modified — this is an append-only audit log.

Columns:
  Date          — ISO date of the pipeline run (YYYY-MM-DD)
  Run Timestamp — Full UTC timestamp of this run
  Title         — Article or post title
  Source        — Feed name or subreddit (r/name)
  URL           — Canonical URL
  Theme         — AI-detected career theme
  Score         — Relevance score (1-10)
  Audience      — Audience segment (e.g. mid_career_professional)
  Coaching Flag — TRUE/FALSE
  Canada        — TRUE/FALSE (Canada-relevant signal)
  Sentiment     — positive / neutral / negative
  Urgency       — high / medium / low
  Summary       — One-paragraph AI summary
  Power Quote   — Most emotionally resonant quote
  Run ID        — Identifies which pipeline run produced this row
"""

import logging
import hashlib
from datetime import datetime, timezone
from pathlib import Path

log = logging.getLogger("excel_logger")

ROOT = Path(__file__).parent.parent
LOG_PATH = ROOT / "data" / "intelligence_log.xlsx"

COLUMNS = [
    "Date",
    "Run Timestamp",
    "Title",
    "Source",
    "URL",
    "Theme",
    "Score",
    "Audience",
    "Coaching Flag",
    "Canada",
    "Sentiment",
    "Urgency",
    "Summary",
    "Power Quote",
    "Run ID",
]

# Column widths (characters) for readability
COLUMN_WIDTHS = {
    "Date": 12,
    "Run Timestamp": 22,
    "Title": 70,
    "Source": 28,
    "URL": 60,
    "Theme": 28,
    "Score": 7,
    "Audience": 28,
    "Coaching Flag": 14,
    "Canada": 8,
    "Sentiment": 12,
    "Urgency": 10,
    "Summary": 90,
    "Power Quote": 80,
    "Run ID": 18,
}


def _run_id():
    """Generate a short unique run ID based on current UTC time."""
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _item_to_row(item, run_date, run_timestamp, run_id):
    """Convert one enriched item dict to an Excel row list."""
    enrichment = item.get("enrichment", {})
    subreddit = item.get("subreddit", "")
    feed_name = item.get("feed_name", "")
    source_display = f"r/{subreddit}" if subreddit else (feed_name or item.get("source", "Unknown"))

    theme = (item.get("theme") or "").replace("_", " ").title()
    audience = (item.get("audience_segment") or "").replace("_", " ").title()
    sentiment = (enrichment.get("sentiment") or "").title()
    urgency = (enrichment.get("urgency") or "").title()
    summary = enrichment.get("summary", "") or ""
    power_quote = item.get("power_quote", "") or ""

    canada_flag = bool(
        enrichment.get("canada_relevant") or
        subreddit in {"canada", "toronto", "ontario", "torontoJobs"}
    )

    return [
        run_date,
        run_timestamp,
        item.get("title", "")[:200],          # cap very long titles
        source_display[:80],
        item.get("url", "")[:500],
        theme[:60],
        item.get("relevance_score", 0),
        audience[:60],
        item.get("coaching_flag", False),
        canada_flag,
        sentiment[:20],
        urgency[:20],
        summary[:600],
        power_quote[:400],
        run_id,
    ]


def _apply_header_style(ws, header_fill, header_font, header_alignment):
    """Apply formatting to header row."""
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def _get_existing_urls(ws):
    """Return set of all URLs already in the sheet (column E = index 5)."""
    existing = set()
    url_col_idx = COLUMNS.index("URL") + 1  # 1-based
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and len(row) >= url_col_idx:
            url = row[url_col_idx - 1]
            if url:
                existing.add(str(url).strip())
    return existing


def append_to_excel(enriched_items, run_date=None, run_timestamp=None):
    """
    Append new enriched items to the master Excel log.
    Deduplicates by URL — existing URLs are skipped.

    Args:
        enriched_items: list of enriched item dicts from claude_enricher
        run_date:       override date string (YYYY-MM-DD); defaults to today
        run_timestamp:  override ISO timestamp; defaults to now UTC

    Returns:
        dict with keys: appended, skipped_duplicate, file_path
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        log.error("openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        return {"appended": 0, "skipped_duplicate": 0, "file_path": str(LOG_PATH), "error": "openpyxl missing"}

    if run_date is None:
        run_date = datetime.now().strftime("%Y-%m-%d")
    if run_timestamp is None:
        run_timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    run_id = _run_id()

    # ── Load or create workbook ───────────────────────────────────────────────
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    if LOG_PATH.exists():
        wb = openpyxl.load_workbook(str(LOG_PATH))
        ws = wb.active
        log.info(f"Loaded existing Excel log: {LOG_PATH} ({ws.max_row - 1} existing rows)")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Intelligence Log"

        # Write header
        ws.append(COLUMNS)

        # Style header row
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        _apply_header_style(ws, header_fill, header_font, header_alignment)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Set column widths
        for col_idx, col_name in enumerate(COLUMNS, 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = COLUMN_WIDTHS.get(col_name, 20)

        log.info(f"Created new Excel log: {LOG_PATH}")

    # ── Deduplicate ───────────────────────────────────────────────────────────
    existing_urls = _get_existing_urls(ws)
    log.info(f"Found {len(existing_urls)} existing URLs in log (deduplication)")

    # ── Append new rows ───────────────────────────────────────────────────────
    appended = 0
    skipped = 0

    # Style for alternating rows
    try:
        from openpyxl.styles import PatternFill
        alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
    except Exception:
        alt_fill = None

    for item in enriched_items:
        url = item.get("url", "").strip()
        if url and url in existing_urls:
            skipped += 1
            continue

        row_data = _item_to_row(item, run_date, run_timestamp, run_id)
        ws.append(row_data)

        # Alternate row shading
        if alt_fill and ws.max_row % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

        # Wrap text in Summary and Title columns
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("Summary", "Title", "Power Quote")))

        if url:
            existing_urls.add(url)
        appended += 1

    # ── Auto-filter on header ─────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Save ─────────────────────────────────────────────────────────────────
    wb.save(str(LOG_PATH))

    log.info(
        f"Excel log updated: {appended} new rows appended, "
        f"{skipped} duplicates skipped → {LOG_PATH}"
    )

    return {
        "appended": appended,
        "skipped_duplicate": skipped,
        "file_path": str(LOG_PATH),
        "run_id": run_id,
    }


def get_log_summary():
    """
    Return a quick summary of the Excel log (row count, date range, top themes).
    Useful for health checks.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed"}

    if not LOG_PATH.exists():
        return {"error": "Log file does not exist yet", "path": str(LOG_PATH)}

    wb = openpyxl.load_workbook(str(LOG_PATH), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        return {"total_rows": 0, "file_path": str(LOG_PATH)}

    dates = [r[0] for r in rows if r[0]]
    themes = [r[5] for r in rows if r[5]]
    coaching_flags = [r[8] for r in rows if r[8] is True]

    from collections import Counter
    top_themes = Counter(themes).most_common(5)

    return {
        "total_rows": len(rows),
        "date_range": f"{min(dates)} to {max(dates)}" if dates else "N/A",
        "coaching_flags": len(coaching_flags),
        "top_themes": top_themes,
        "file_path": str(LOG_PATH),
    }


if __name__ == "__main__":
    # Quick test / summary print
    summary = get_log_summary()
    print("Excel Log Summary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")
